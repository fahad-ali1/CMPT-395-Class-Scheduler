"""
Author: Mike Lee, Sankalp Shrivastav
ID: 3067768, 3106374
Date: 20/03/2023
Purpose: Collection of classes and methods needed for the project
"""

import re

from openpyxl.reader.excel import load_workbook


class Course:
    """
    Course Class:
    Represents a course
    Attributes include:
    courseName = name of the course
    courseDescript = course description
    lectureLength = length of each lecture
    courseType = the type of each course(normal, online/virtual, lab)
    schedulingInstructions a string indicating how to schedule a course, for example: NBOA = do not schedule before or
    after another course
    numberOfSessions = a string indicating n number of sessions. getNumOfSessions will return just the number
    """
    def __init__(self, courseName="", courseDescript="", totalTranscriptHours=0):
        self.courseName = courseName
        self.courseDescript = courseDescript
        self.totalTranscriptHours = totalTranscriptHours
        self.lectureLength = 0.0
        self.numberOfSessions = 0
        self.courseType = ""
        self.schedulingInstructions = ""

    def setCourseName(self, courseName): #change course name
        self.courseName = courseName

    def setCourseDes(self, courseDes): #add a description
        self.courseDescript = courseDes

    def setNumberOfSessions(self, sessions): #number of sessions
        self.numberOfSessions = sessions

    def setTranscriptHours(self, transcriptHours): #total class time
        self.totalTranscriptHours = transcriptHours

    def setCourseType(self, status): 
        self.courseType = status

    def setLectureLength(self, lecLen): # time block for each lecture
        self.lectureLength = lecLen

    def setschedulingInstructions(self, instructions):
        self.schedulingInstructions = instructions

    def getschedulingInstructions(self):
        return self.schedulingInstructions

    def getcourseName(self):
        return self.courseName

    def getdescription(self):
        return self.courseDescript

    def getcourseType(self):
        return self.courseType

    def getLectureLength(self):
        return self.lectureLength

    def getNumOfSessions(self):
        splitStr = str(self.numberOfSessions).split(" ")
        numVal = splitStr[0]
        return numVal

    def printCourseDetails(self):
        print(self.courseName, " ", self.courseDescript, " ", self.totalTranscriptHours)


class Program:

    """
    Program: Represents a program at macewan
        i.e.: BCOM, Bachelor of something
    Attributes include:
    programType = the type(name) of the program

    three academic terms
    term1 = list of all courses(course class) for term 1
    term2 = list of all courses(course class) for term 2
    term3 = list of all courses(course class) for term 3
    """
    def __init__(self, programType=""):
        self.programType = programType
        self.term1 = []
        self.term2 = []
        self.term3 = []

    def setProgram(self, programName):
        self.programType = programName

    def addToTerm(self, course, termNum):
        termNum = termNum.strip()
        if termNum == "Term 1":
            self.term1.append(course)
        elif termNum == "Term 2":
            self.term2.append(course)
        elif termNum == "Term 3":
            self.term3.append(course)

    def addProgramType(self, programType):
        self.programType = programType

    def printProgram(self):
        print(self.programType)

    def getAllTerms(self):
        allTerms = [self.term1, self.term2, self.term3]
        return allTerms

class Classroom:
    """
    Classroom:
    Represents a physical classroom
    Attributes include:
    classRoomNumber = classroom number/name
    normalCapacity = capacity
    lab = is it a lab
    isGhost = is it a ghost room (room that would be needed for extra students)
    schedule = schedule linked list representing the schedule
    """
    def __init__(self, classRoomNumber='', normalCapacity=0, lab=False):
        self.classRoomNumber = classRoomNumber
        self.normalCapacity = normalCapacity
        self.lab = lab
        self.isGhost = False
        self.schedule = None
        self.currentStudents = 0 #will make a setter for it
        self.inUse = False

    def setGhost(self):
        self.isGhost = True

    def setSchedule(self, schedule):
        self.schedule = schedule

    def printClassroom(self):
        print("Classroom #: ", self.classRoomNumber, "Normal Capacity: ", 
              self.normalCapacity, "lab: ", self.lab, " isGhost: ", self.isGhost)

class Cohort:
    """
    Cohort Class takes in a cohort name and size initially
    Methods will be created to add a programs to mainProgramCourses
    and electiveProgramCourses
    """
    def __init__(self, cohortName="", size=0):
        self.cohortName = cohortName
        self.size = size
        self.mainProgramCourses = []  # program class
        self.electiveProgramCourses = []  # program class
        self.prereq = {}

    def getCohortName(self):
        return self.cohortName
    
    # need to tuch base on default size
    def updateCohortSize(self, size=0):
        self.size = size

    #type = M:main, E:elective, P:preReq
    def addProgCourse(self, type="M"):

        pass

    def displayCohort(self, type="M"):

        requirnments = self.mainProgramCourses
        if type is not "M": requirnments = self.electiveProgramCourses

        print(  f"Name: {self.cohortName}\n"\
                f"Size: {self.size}\n"\
                f"Requirnments: {requirnments}\n"\
                f"Pre-Reqs: {self.prereq}")

# ScheduleLinkedList:
# Represents a schedule for a classroom object
# Attributes include:
#time: time uses 2400 time format and starts at 8am. Each time a node is added the time will increase by that nodes
#lecture length. The time starts at 8am and should not exceed 4pm
#cohort: represents a cohort
#startDate: represents the courses start date
#endDate: repsrents the courses end date
# head = head node of schedule (doubly linked list)
class ScheduleLinkedList:
    def __init__(self, totalHours=800):
        self.totalHours = totalHours
        self.head = None

    def addNodePush(self, time, cohort, startDate, endDate):
        newNode = scheduleNode(time, cohort, startDate, endDate)
        newNode.next = self.head
        if self.head is not None:
            self.head.prev = newNode
        self.head = newNode
    
    def getCohort(self, node):

        # if is empty return none
        if self.head is None: return None

        cohortName = node.cohortName # name of the searching cohort
        currentCohort = self.head # first item

        while True: # while not at the end keep going

            # currentCohort.cohort.displayCohort()

            if currentCohort.cohort.cohortName is cohortName: # FOUND!!!
                
                # add time management for returning it
                return currentCohort.cohort.cohortName, currentCohort.startHour, currentCohort.endHour

            currentCohort = currentCohort.next # swap to next node

            if currentCohort is None: break

        return None # when not found

     #checks to see if the schedule is full
    #full = True; not full = False
    def checkIfScheduleFull(self):
        if self.totalHours == 1600: return True
        return False

    def updateTime(self, node):
        hours = self.totalHours
        if node is not None:
            if (node.time * 100) + hours > 1600:
                return -1 # -1 is an error indicating full
            else:
                self.totalHours = hours + (node.time * 100)
        else:
            return hours
        return int(self.totalHours)

    def getTotalHours(self):
        return int(self.totalHours)

# ScheduleNode:
# Represents a lecture time block for a classroom's schedule
# Attributes include:
# time = lecture length
# cohort = cohort
# startDate = course scheduled start date
# endDate = course scheduled end date
# prev = prev lecture block (node)
# next = next lecture block (node)
class scheduleNode:
    def __init__(self, time, cohort, startHour, endHour):
        self.time = time #time should be 1.5 or 2.0, etc..
        self.cohort = cohort
        self.startHour = startHour
        self.endHour = endHour
        self.startDate = 0
        self.endDate = 0
        self.prev = None
        self.next = None

    #input: Convert the lecture time from a course object to it's time in hours.
    #output: return the hour
    def convertLectureToHour(self):
        return self.time * 100

    def setStartDate(self, StartDate):
        self.startDate = StartDate

    def setEndDate(self, EndDate):
        self.endDate = EndDate


#input: Convert the lecture time from a course object to it's time in hours.
#output: return the hour
def convertLectureToHour(course):
    return course.lectureLength * 100
#need a function to get start and end dates in 2400 hour format
# start/end is for schedule node

# getClassrooms:
# Helper Fucntion
# Purpose: Used to gather list of classroom objects representing all available classrooms
def getClassrooms():
    wb = load_workbook('AllCourses.xlsx')
    allClassrooms = []
    ws = wb.worksheets[8]
    for i in range(2, ws.max_row + 1):
        if "Lab" in ws['A' + str(i)].value:
            classroom = Classroom(ws['A' + str(i)].value, ws['B' + str(i)].value, True)
            allClassrooms.append(classroom)
        else:
            classroom = Classroom(ws['A' + str(i)].value, ws['B' + str(i)].value)
            allClassrooms.append(classroom)
    return allClassrooms

# getAllPrograms
# Helper Function
# Purpose: Used to gather a list of program objects representing all available programs at macewan
def getAllPrograms():
    wb = load_workbook('AllCourses.xlsx')
    term = None
    allPrograms = []
    for i in range(0, 8):
        ws1 = wb.worksheets[i]
        program = Program(programType=ws1.title)
        for j in range(2, ws1.max_row + 1):
            if re.search('Term [1-3]', ws1['A' + str(j)].value):
                term = ws1['A' + str(j)].value
                pass

            elif ws1['B' + str(j)].value is not None:
                course = Course(courseName=ws1['A' + str(j)].value, courseDescript=ws1['B' + str(j)].value,
                                totalTranscriptHours=ws1['C' + str(j)].value)
                course.setCourseType(str(checkIfLab(ws1['E' + str(j)].value)))
                course.setLectureLength(float(getMinimumTime(ws1['F' + str(j)].value)))
                if ws1['G' + str(j)].value is None:
                    course.setNumberOfSessions("")
                else:
                    course.setNumberOfSessions(ws1['G' + str(j)].value)

                program.addToTerm(course, term)
        allPrograms.append(program)
    return allPrograms

def calculateSessions(transcriptHours, lectureLen):
    pass

# getProgramNumbers
# Helper Function
#
def getProgramNumbers(fileName):
    try:
        wb = load_workbook(fileName)
        # wb = load_workbook("semester_data.xlsx") #fast testing
        ws = wb.active
        programNumbersList = []
        for i in range(2, ws.max_row + 1):
            programNumbersList.append([ws['B' + str(i)].value, ws['C' + str(i)].value, ws['D' + str(i)].value])
        return programNumbersList
    except:
        print("Error: File not found")
        return -1

# Expected input: String or None type value from cell F of Allcourse.xlsx
# Expected output: If None type is found then return float indicating default lecture length is returned (1.5 hours)
# Else if value is found then return float indicating minimum lecture length
def getMinimumTime(data):
    if data is None:
        return 1.5
    else:
        return float(data[0])


# Expected input: value from cell E of AllCourse.xlsx
# Expected output: returns a string indicating the lab status of a course
def checkIfLab(data):
    if data is None:
        return "Normal"
    elif data == "Lab":
        return "Lab"
    elif data == "Both":
        return "Both"
    elif data == "Virtual":
        return "Virtual"
    elif data == "Online":
        return "Online"

#This function is a help function designed to help with assigning number of sessions to a course
#If the number of course sessions do not divide evenly with the total transcript hours then add an extra hour
#Else if evenly divided, return the number without any change.
#Expected input: a number either an integer or a float
#Expected output returns a number based on if the input had a decimal or not. If no decimal then return number unchanged
#Else if input has a decimal, then return integer value + 1 of input
def checkDecimal(num):
    numToStr = str(num)
    if numToStr.find(".") < 1:
        return int(num) + 1
    else:
        return num


def getCourse(fileName, courseName):

    wb = load_workbook(fileName)

    for sheet in range(8):

        ws = wb.worksheets[sheet]
        
        for row in range(2, ws.max_row + 1):

            #look for the course name if found return the location [row, col]

            if re.search("^Term", ws["A" + str(row)].value): continue
            # print(ws["A" + str(row)].value)
            if re.search(courseName, ws["A" + str(row)].value): return [sheet, row]

    return None

def changeCourseInfo(courseLocation, courseInfo, fileName):

    sheet, cRow = courseLocation[0], courseLocation[1]
    wb = load_workbook(fileName)

    ws = wb.worksheets[sheet]
    count = 0
    
    for cells in ws[str(cRow)]:
        #swap the info from courseInfo

        cells.value = courseInfo[count]
        # print(cells.value, end="\t") #comment this when demo is done
        count += 1
    wb.save(fileName) #saving the edited file as 


'''
Legend for excel file codes:
SA = Schedule after all classes are done, end of the term
NBOA = Do not schedule immediately before or after other courses
Twice/Half = Class should be schedule twice a week half way through the term
'''
'''
ALLCOURSES.XLSX LEGEND:
A:
Term/course name
B:
Course Description
C:
Total Transcript Hours
D:
Prereqs
E:
Course Type(lab, normal, online/virtual, etc...)
F:
Timeslot
G:
Number of sessions
H:
Order classes should be scheduled(Twice a week halfway through a semester, etc..)
'''
'''
Hours   # of sessiosn
15      10
21      14
35      24
50      34

WHAT TO WORK ON:
return type = string cohort; float start hour; float end hour;
Ideas:
'''
def returnStuff(data):
    pass