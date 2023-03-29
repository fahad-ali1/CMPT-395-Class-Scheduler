'''
Author: Fahad Ali
Description: Create main GUI app, with multiple functionalities
'''

import sys, time, openpyxl, datetime, time, calendar, random
from openpyxl.styles import PatternFill, Border, Side
from PyQt5.QtWidgets import *
from PyQt5.QtCore import *
from PyQt5 import QtGui, uic
from itertools import chain

# Import local code
from lib.fileio import getProgramNumbers, getClassrooms
from lib.cohorts import Students 
from lib.scheduler import Scheduler, Week, Day, timeBlock

class MainMenu(QWidget):
    def __init__(self):
        super(MainMenu, self).__init__()

        # Load the UI file
        uic.loadUi("MainWindow.ui", self)

        # call students class from students module
        self._studentCohort = Students()

        # list to zero all the spin boxes quickly
        self.zero = [0,0,0,0,0,0,0,0]

        #------------------------- Inputs Tab-----------------------------
        
        # Define student inputs 
        self.terminput = self.findChild(QSpinBox, "TermInput")
        self.BCOMinput = self.findChild(QSpinBox, "BCOMinput")
        self.PCOMinput = self.findChild(QSpinBox, "PCOMinput")
        self.PMinput = self.findChild(QSpinBox, "PMinput")
        self.BAinput = self.findChild(QSpinBox, "BAinput")
        self.SCMTinput = self.findChild(QSpinBox, "SCMTinput")
        self.BKinput = self.findChild(QSpinBox, "BKinput")
        self.FSinput = self.findChild(QSpinBox, "FSinput")
        self.DXDinput = self.findChild(QSpinBox, "DXDinput")
        
        # Define buttons
        self.createCohort = self.findChild(QPushButton, "CreateCohort")
        self.revertButton = self.findChild(QPushButton, "RevertButton")
        self.revertButton2 = self.findChild(QPushButton, "RevertButton2")
        self.revertButton3 = self.findChild(QPushButton, "RevertButton3")

        self.uploadButton = self.findChild(QPushButton, "UploadFileButton")
        self.roomsButton = self.findChild(QPushButton, "RoomsButton")
        self.uploadButton2 = self.findChild(QPushButton, "UploadFileButton2")


        #------------------------- Progress bars ------------------------------
        self.progressBarUpload = self.findChild(QProgressBar, "ProgressBarUpload")
        self.progressBarUpload2 = self.findChild(QProgressBar, "ProgressBarUpload2")
        self.progressBarUpload3 = self.findChild(QProgressBar, "ProgressBarUpload3")

        self.progressBarDownload = self.findChild(QProgressBar, "ProgressBarDownload")
        self.progressBarSaveAllCohort = self.findChild(QProgressBar,
                                                       "ProgressBarSaveAllCohort")

        self.progressBarSaveSchedule = self.findChild(QProgressBar, "ScheduleSaveBar")

        #-------------------------  Cohort Tab --------------------------------
        # Set cohort tables
        self.cohortTable1 = self.findChild(QTableWidget, "CohortTable1")
        self.cohortTable2 = self.findChild(QTableWidget, "CohortTable2")
        self.cohortTable3 = self.findChild(QTableWidget, "CohortTable3")

        #------------------------- Schedules Tab -----------------------------
        self.roomNums = []
        
        # Set schedule tables per room   
        self.scheduleTab = self.findChild(QTabWidget, "ScheduleTables")

        self.room1 = self.findChild(QTableWidget, "ScheduleR1")
        self.room1week = self.findChild(QLabel, "WeekNumR1")
        self.room2 = self.findChild(QTableWidget, "ScheduleR2")
        self.room2week = self.findChild(QLabel, "WeekNumR2")
        self.room3 = self.findChild(QTableWidget, "ScheduleR3")
        self.room3week = self.findChild(QLabel, "WeekNumR3")
        self.room4 = self.findChild(QTableWidget, "ScheduleR4")
        self.room4week = self.findChild(QLabel, "WeekNumR4")
        self.room5 = self.findChild(QTableWidget, "ScheduleR5")
        self.room5week = self.findChild(QLabel, "WeekNumR5")
        self.room6 = self.findChild(QTableWidget, "ScheduleR6")
        self.room6week = self.findChild(QLabel, "WeekNumR6")
        self.room7 = self.findChild(QTableWidget, "ScheduleR7")
        self.room7week = self.findChild(QLabel, "WeekNumR7")
        self.room8 = self.findChild(QTableWidget, "ScheduleR8")
        self.room8week = self.findChild(QLabel, "WeekNumR8")
        self.room9 = self.findChild(QTableWidget, "ScheduleR9")
        self.room9week = self.findChild(QLabel, "WeekNumR9")
        self.room10 = self.findChild(QTableWidget, "ScheduleR10")
        self.room10week = self.findChild(QLabel, "WeekNumR10")
        self.room11 = self.findChild(QTableWidget, "ScheduleR11")
        self.room11week = self.findChild(QLabel, "WeekNumR11")
        self.room12 = self.findChild(QTableWidget, "ScheduleR12")
        self.room12week = self.findChild(QLabel, "WeekNumR12")
        self.room13 = self.findChild(QTableWidget, "ScheduleR13")
        self.room13week = self.findChild(QLabel, "WeekNumR13")
        self.room14 = self.findChild(QTableWidget, "ScheduleR14")
        self.room14week = self.findChild(QLabel, "WeekNumR14")
        self.room15 = self.findChild(QTableWidget, "ScheduleR15")
        self.room15week = self.findChild(QLabel, "WeekNumR15")
        self.room16 = self.findChild(QTableWidget, "ScheduleR16")
        self.room16week = self.findChild(QLabel, "WeekNumR16")
        self.room17 = self.findChild(QTableWidget, "ScheduleR17")
        self.room17week = self.findChild(QLabel, "WeekNumR17")
        self.room18 = self.findChild(QTableWidget, "ScheduleR18")
        self.room18week = self.findChild(QLabel, "WeekNumR18")
        self.room19 = self.findChild(QTableWidget, "ScheduleR19")
        self.room19week = self.findChild(QLabel, "WeekNumR19")
        self.room20 = self.findChild(QTableWidget, "ScheduleR20")
        self.room20week = self.findChild(QLabel, "WeekNumR20")
        
        # buttons
        self.room1prev = self.findChild(QPushButton, "PrevWeekR1")
        self.room1next = self.findChild(QPushButton, "NextWeekR1")
        self.room2prev = self.findChild(QPushButton, "PrevWeekR2")
        self.room2next = self.findChild(QPushButton, "NextWeekR2")
        self.room3prev = self.findChild(QPushButton, "PrevWeekR3")
        self.room3next = self.findChild(QPushButton, "NextWeekR3")
        self.room4prev = self.findChild(QPushButton, "PrevWeekR4")
        self.room4next = self.findChild(QPushButton, "NextWeekR4")
        self.room5prev = self.findChild(QPushButton, "PrevWeekR5")
        self.room5next = self.findChild(QPushButton, "NextWeekR5")
        self.room6prev = self.findChild(QPushButton, "PrevWeekR6")
        self.room6next = self.findChild(QPushButton, "NextWeekR6")
        self.room7prev = self.findChild(QPushButton, "PrevWeekR7")
        self.room7next = self.findChild(QPushButton, "NextWeekR7")
        self.room8prev = self.findChild(QPushButton, "PrevWeekR8")
        self.room8next = self.findChild(QPushButton, "NextWeekR8")
        self.room9prev = self.findChild(QPushButton, "PrevWeekR9")
        self.room9next = self.findChild(QPushButton, "NextWeekR9")
        self.room10prev = self.findChild(QPushButton, "PrevWeekR10")
        self.room10next = self.findChild(QPushButton, "NextWeekR10")
        self.room11prev = self.findChild(QPushButton, "PrevWeekR11")
        self.room11next = self.findChild(QPushButton, "NextWeekR11")
        self.room12prev = self.findChild(QPushButton, "PrevWeekR12")
        self.room12next = self.findChild(QPushButton, "NextWeekR12")
        self.room13prev = self.findChild(QPushButton, "PrevWeekR13")
        self.room13next = self.findChild(QPushButton, "NextWeekR13")
        self.room14prev = self.findChild(QPushButton, "PrevWeekR14")
        self.room14next = self.findChild(QPushButton, "NextWeekR14")
        self.room15prev = self.findChild(QPushButton, "PrevWeekR15")
        self.room15next = self.findChild(QPushButton, "NextWeekR15")
        self.room16prev = self.findChild(QPushButton, "PrevWeekR16")
        self.room16next = self.findChild(QPushButton, "NextWeekR16")
        self.room17prev = self.findChild(QPushButton, "PrevWeekR17")
        self.room17next = self.findChild(QPushButton, "NextWeekR17")
        self.room18prev = self.findChild(QPushButton, "PrevWeekR18")
        self.room18next = self.findChild(QPushButton, "NextWeekR18")
        self.room19prev = self.findChild(QPushButton, "PrevWeekR19")
        self.room19next = self.findChild(QPushButton, "NextWeekR19")
        self.room20prev = self.findChild(QPushButton, "PrevWeekR20")
        self.room20next = self.findChild(QPushButton, "NextWeekR20")


        # List of schedules per room
        self.rooms = [self.room1, self.room2, self.room3, self.room4,\
            self.room5, self.room6, self.room7, self.room8, self.room9, \
                self.room10, self.room11, self.room12, self.room13, \
                    self.room14, self.room15, self.room16, self.room17,\
                        self.room18, self.room19, self.room20]
        
        # List of week number for each room tab
        self.weekNum = [self.room1week, self.room2week, self.room3week, \
            self.room4week, self.room5week, self.room6week, self.room7week,\
                self.room8week, self.room9week, self.room10week, \
                    self.room11week, self.room12week, self.room12week, \
                        self.room12week, self.room13week, self.room14week, \
                            self.room15week, self.room16week, self.room17week,\
                                self.room18week, self.room19week, \
                                 self.room20week]

        # call method and create schedules for all rooms
        self.daysIterator = 0
        self.week = 1
        self.weeksInYear = []
        self.year = datetime.date.today().year

        # format calender to year
        self.format_calendar(self.year)
        self.create_dates()
        
        # create proper room names and enable the correct room
        self.enable_room()       

        #------------- Buttons for Saving and Downloading template ------------

        # Define buttons for schedule saving
        self.saveCurrent1 = self.findChild(QPushButton, "SaveCurrentSchedule1")
        self.saveAllCohort = self.findChild(QPushButton, "SaveAllCohortButton")
        self.saveAllSchedule = self.findChild(QPushButton, "SaveAllScheduleButton")

        self.downloadTemplate = self.findChild(QPushButton, "DownloadTemplate")

        #------------------------- Connect Buttons -----------------------------

        # Call revert_changes function 
        self.revertButton.clicked.connect(self.revert_changes)
        self.revertButton2.clicked.connect(self.revert_changes)
        self.revertButton3.clicked.connect(self.revert_changes)

        # Open file browser if upload file clicked
        self.uploadButton.clicked.connect(self.upload_file)
        self.uploadButton2.clicked.connect(self.upload_file)
        self.roomsButton.clicked.connect(self.upload_room_file)

        # Create cohorts when button clicked
        self.createCohort.clicked.connect(self.create_cohorts)
        self.createCohort.clicked.connect(self.add_cohort_table)

        # Download/save template when clicked
        self.downloadTemplate.clicked.connect(self.download_template)

        # Download/save all cohorts as .xlsx
        self.saveAllCohort.clicked.connect(self.save_all_cohort)

        # Download/save all schedules as .xlsx
        self.saveAllSchedule.clicked.connect(self.save_all_schedule)
        
        self.cohortUploaded = False
        
        # Create next week schedule
        self.room1next.clicked.connect(self.week_next)
        self.room2next.clicked.connect(self.week_next)
        self.room3next.clicked.connect(self.week_next)
        self.room4next.clicked.connect(self.week_next)
        self.room5next.clicked.connect(self.week_next)
        self.room6next.clicked.connect(self.week_next)
        self.room7next.clicked.connect(self.week_next)
        self.room8next.clicked.connect(self.week_next)
        self.room9next.clicked.connect(self.week_next)
        self.room10next.clicked.connect(self.week_next)
        self.room11next.clicked.connect(self.week_next)
        self.room12next.clicked.connect(self.week_next)
        self.room13next.clicked.connect(self.week_next)
        self.room14next.clicked.connect(self.week_next)
        self.room15next.clicked.connect(self.week_next)
        self.room16next.clicked.connect(self.week_next)
        self.room17next.clicked.connect(self.week_next)
        self.room18next.clicked.connect(self.week_next)
        self.room19next.clicked.connect(self.week_next)
        self.room20next.clicked.connect(self.week_next)
        
        # Create previous week schedule
        self.room1prev.clicked.connect(self.week_prev)
        self.room2prev.clicked.connect(self.week_prev)
        self.room3prev.clicked.connect(self.week_prev)
        self.room4prev.clicked.connect(self.week_prev)
        self.room5prev.clicked.connect(self.week_prev)
        self.room6prev.clicked.connect(self.week_prev)
        self.room7prev.clicked.connect(self.week_prev)
        self.room8prev.clicked.connect(self.week_prev)
        self.room9prev.clicked.connect(self.week_prev)
        self.room10prev.clicked.connect(self.week_prev)
        self.room11prev.clicked.connect(self.week_prev)
        self.room12prev.clicked.connect(self.week_prev)
        self.room13prev.clicked.connect(self.week_prev)
        self.room14prev.clicked.connect(self.week_prev)
        self.room15prev.clicked.connect(self.week_prev)
        self.room16prev.clicked.connect(self.week_prev)
        self.room17prev.clicked.connect(self.week_prev)
        self.room18prev.clicked.connect(self.week_prev)
        self.room19prev.clicked.connect(self.week_prev)
        self.room20prev.clicked.connect(self.week_prev)

    def spin_box_values(self, value, tables):
            '''
            Description: changes spinbox values
            '''
            # set value of each box to a certain value
            self.terminput.setValue(tables)
            self.BCOMinput.setValue(value[0])
            self.PCOMinput.setValue(value[1])
            self.PMinput.setValue(value[2])
            self.BAinput.setValue(value[3])
            self.SCMTinput.setValue(value[4])
            self.BKinput.setValue(value[5])
            self.FSinput.setValue(value[6])
            self.DXDinput.setValue(value[7])

    def revert_changes(self):
        '''
        Description: revert all changes
        '''
        for tables in range (1,4):
            # Set all spin box objects to 0
            self.spin_box_values(self.zero, tables)

            # Delete all cohorts and information on the cohort tables
            self.create_cohorts()
            self.add_cohort_table()

        # return to default values
        self.terminput.setValue(1)

        self.progressBarUpload.setValue(0)
        self.progressBarUpload2.setValue(0)
        self.progressBarUpload3.setValue(0)
        
        self.progressBarUpload.setFormat("")
        self.progressBarUpload2.setFormat("")
        self.progressBarUpload3.setFormat("")
        
        self.progressBarUpload.setTextVisible(False)
        self.progressBarUpload2.setTextVisible(False)
        self.progressBarUpload3.setTextVisible(False)
        
        self.progressBarDownload.setValue(0)
        self.progressBarDownload.setFormat("")
        self.progressBarDownload.setTextVisible(False)

        self.progressBarSaveAllCohort.setValue(0)
        self.progressBarSaveAllCohort.setFormat("")
        self.progressBarSaveAllCohort.setTextVisible(False)

        self.progressBarSaveSchedule.setValue(0)
        self.progressBarSaveSchedule.setFormat("")
        self.progressBarSaveSchedule.setTextVisible(False)

    def create_cohorts(self):
        '''
        Description: assign input to students class to create cohorts
        '''
        # assign input to class attributes
        self._studentCohort._term = int(self.terminput.text())
        self._studentCohort._BCOMStudents = int(self.BCOMinput.text())
        self._studentCohort._PCOMStudents = int(self.PCOMinput.text())
        self._studentCohort._PMStudents = int(self.PMinput.text())
        self._studentCohort._BAStudents = int(self.BAinput.text())
        self._studentCohort._SCMTStudents = int(self.SCMTinput.text())
        self._studentCohort._BKStudents = int(self.BKinput.text())
        self._studentCohort._FSStudents = int(self.FSinput.text())
        self._studentCohort._DXDStudents = int(self.DXDinput.text())

        # create cohort final attribute
        self._cohortFinal = self._studentCohort.cohorts_final()
        self._flattened_cohorts = list(chain.from_iterable(self._cohortFinal))
        self.create_schedule()     
            
    def add_cohort_table(self):
        '''
        Description: create cohort tables
        '''
        # assign variables
        self.tables = [self.cohortTable1, self.cohortTable2, self.cohortTable3]
        self.tables[int(self.terminput.text())-1].\
            setRowCount(len(max(self._cohortFinal,key=len)))
            
        j = 0
        # create cells for each table and add information
        for cohortLists in self._cohortFinal:
            color = self.random_background_color()
            i = 0
            for cohorts in cohortLists:
                self.tables[int(self.terminput.text())-1]\
                    .setItem(i, j, QTableWidgetItem(str(cohorts)))
                self.tables[int(self.terminput.text())-1]\
                    .item(i, j).setBackground(QtGui.QColor(color[0],color[1],color[2]))
                i += 1
            j += 1

    def upload_file(self):
        '''
        Description: open dialog box to select file and then auto input all
        students into cohorts and create schedule for each room
        '''
        # open file browser to choose file
        filter = ".xlsx(*.xlsx)"
        path = QFileDialog.getOpenFileName(self, 'Open File','', filter)
        # give path if the user does not hit cancel 
        if path != ("", ""):
            filename = path[0]
            # input file and extract student inputs
            fileinput = getProgramNumbers(filename)

            studentsByProgarm = {"BCOM":fileinput[0], "PCOM":fileinput[1],
                                "PM":fileinput[2], "BA":fileinput[3],
                                "SCMT":fileinput[4], "BK":fileinput[5],
                                "FS":fileinput[6], "DXD":fileinput[7]}
            term1, term2, term3 = [], [], []
            inputs = [term1, term2, term3]
            # loop through dictionary values
            for studentInputs in studentsByProgarm.values():
                # studentInputs[0] = term 1
                # studentInputs[1] = term 2
                # studentInputs[2] = term 3
                term1.append(int(studentInputs[0]))
                term2.append(int(studentInputs[1]))
                term3.append(int(studentInputs[2]))

            # assign and create cohorts in the tables from the file 
            for tables in range (1,4):
                self.spin_box_values(inputs[tables-1], tables)

                self.create_cohorts()
                self.add_cohort_table()

            # Set value to progress bar
            self.progressBarUpload.setValue(100)
            self.progressBarUpload2.setValue(100)
            self.progressBarUpload.setFormat('Upload Complete! (Cohort Tab)')
            self.progressBarUpload.setTextVisible(True)
            self.progressBarUpload2.setFormat('Upload Complete!')
            self.progressBarUpload2.setTextVisible(True)

            # set values 
            self.spin_box_values(self.zero, tables)
            self.terminput.setValue(1)
            self.cohortUploaded = True
        
    def upload_room_file(self):
        '''
        Description: open dialog box to select file and then create room tabs
        '''
        # TODO: add functionality ask Mike for his portion of code
        # open file browser to choose file
        filter = ".xlsx(*.xlsx)"
        path = QFileDialog.getOpenFileName(self, 'Open File','', filter)
        # give path if the user does not hit cancel 
        if path != ("", ""):
            filename = path[0]
            
            # Set value to progress bar
            self.progressBarUpload3.setValue(100)
            self.progressBarUpload3.setFormat('Upload Complete!')
            self.progressBarUpload3.setTextVisible(True)
    
    def download_template(self):
        '''
        Description: download student upload file template
        '''
        # create filter and save file dialog
        filter = ".xlsx(*.xlsx)"
        path = QFileDialog.getSaveFileName(self, 'Save File',
                                           'StudentInputTemplate', filter)

        # if user hits cancel it wont raise error/crash program
        if path != ("", ""):
            savepath = path[0]
            # set cell border
            thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))

            # create workbook
            wb = openpyxl.Workbook()
            sheet = wb.active

            # apply background colors and borders
            for row in sheet ["A2":"A9"]:
                for cell in row:
                    cell.fill = PatternFill(start_color='FFFF00',
                                            end_color='FFFF00', fill_type="solid")
                    cell.border = thin_border
            for row in sheet ["B1":"B9"]:
                for cell in row:
                    cell.fill = PatternFill(start_color='FFD580',
                                            end_color='FFD580', fill_type="solid")
                    cell.border = thin_border
            for row in sheet ["C1":"C9"]:
                for cell in row:
                    cell.fill = PatternFill(start_color='81B622',
                                            end_color='81B622', fill_type="solid")
                    cell.border = thin_border
            for row in sheet ["D1":"D9"]:
                for cell in row:
                    cell.fill = PatternFill(start_color='ADD8E6',
                                            end_color='ADD8E6', fill_type="solid")
                    cell.border = thin_border

            # create term cells
            c1 = sheet.cell(row = 1, column = 2)
            c1.value = "Term 1"
            c2 = sheet.cell(row = 1, column = 3)
            c2.value = "Term 2"
            c3 = sheet.cell(row = 1, column = 4)
            c3.value = "Term 3"

            # create progam cells
            r1 = sheet.cell(row = 2, column = 1)
            r1.value = "BCOM"
            r2 = sheet.cell(row = 3, column = 1)
            r2.value = "PCOM"
            r3 = sheet.cell(row = 4, column = 1)
            r3.value = "PM"
            r4 = sheet.cell(row = 5, column = 1)
            r4.value = "BA"
            r5 = sheet.cell(row = 6, column = 1)
            r5.value = "SCMT"
            r6 = sheet.cell(row = 7, column = 1)
            r6.value = "BK"
            r7 = sheet.cell(row = 8, column = 1)
            r7.value = "FS"
            r8 = sheet.cell(row = 9, column = 1)
            r8.value = "DXD"

            # save workbook to path
            wb.save(savepath)

        # Set value to progress bar
        self.progressBarDownload.setValue(100)
        self.progressBarDownload.setFormat("Download complete!")
        self.progressBarDownload.setTextVisible(True)

    def save_all_schedule(self):
        '''
        Description: saves all schedules into one excel sheet
        '''
        # TODO: mostly hardcoded due to lack of backend code
        filter = ".xlsx(*.xlsx)"
        path = QFileDialog.getSaveFileName(self, 'Save File', 'Schedules', filter)
        # if user hits cancel it wont raise error/crash program
        if path != ("", ""):
            savepath = path[0]

            # create workbook
            wb = openpyxl.Workbook()
            
            # create all sheets
            for worksheet in self.roomNums:
                wb.create_sheet(worksheet)

            # save schedules to a .xlsx file
            for i in range(len(self.roomNums)):
                for c in range(self.rooms[i].columnCount()):
                    for r in range(self.rooms[i].rowCount()):
                        # if an item is in the cell
                        if self.rooms[i].item(r,c):
                            value = self.rooms[i].item(r,c).text()
                            wb.get_sheet_by_name(self.roomNums[i])\
                                .cell(row=r+1, column=c+1).value = value
            wb.save(savepath)

            for i in range(101):
                # Set value to progress bar to create a download animation
                time.sleep(0.011)
                self.progressBarSaveSchedule.setTextVisible(True)
                self.progressBarSaveSchedule.setFormat(f"One moment ... %{i}")
                self.progressBarSaveSchedule.setValue(i)
        self.progressBarSaveSchedule.setFormat("Download complete!")

    def save_all_cohort(self):
        '''
        Description: saves all cohorts into one excel sheet
        '''
        filter = ".xlsx(*.xlsx)"
        path = QFileDialog.getSaveFileName(self, 'Save File', 'Cohorts', filter)
        # if user hits cancel it wont raise error/crash program
        if path != ("", ""):
            savepath = path[0]

            # create workbook
            wb = openpyxl.Workbook()

            # create all sheets
            ws = wb.active
            ws.title = "Cohorts Term 1"
            ws2 = wb.create_sheet("Term 2")
            ws2.title = "Cohorts Term 2"
            ws3 = wb.create_sheet("Term 3")
            ws3.title = "Cohorts Term 3"

            wsList = ["Cohorts Term 1", "Cohorts Term 2", "Cohorts Term 3"]

            # save tables to a .xlsx file
            for i in range (3):
                for c in range(self.tables[i].columnCount()):
                    for r in range(self.tables[i].rowCount()):
                        # if an item is in the cell
                        if self.tables[i].item(r,c):
                            cohortValue = self.tables[i].item(r,c).text()
                            ws = wb.get_sheet_by_name(wsList[i])\
                                .cell(row=r+1, column=c+1).value = cohortValue
            wb.save(savepath)

        self.progressBarSaveAllCohort.setValue(100)
        self.progressBarSaveAllCohort.setFormat("Download complete!")
        self.progressBarSaveAllCohort.setTextVisible(True)

    def format_calendar(self, year):
        '''
        Description: formats calender to useable format
        '''        
        # create a calender object
        self.calendar = calendar.Calendar(firstweekday = 0)
        # loop through every month in year (0-11)
        for i in range(12):
            # loop through every week in year (0 - num of weeks)
            for week in self.calendar.yeardatescalendar(year, 12)[0][i]:
                # loop through each day in a week (0-6)
                for day in range(7):
                    # NOTE: for some reason duplicates exist, so this ensures 
                    # duplicates are removed 
                    if week[day].strftime("%x") not in self.weeksInYear:
                        self.weeksInYear.append(week[day].strftime("%x"))
        
    def create_dates(self):
        '''
        Description: set date into every room
        '''
        # loop through every room
        i = 0
        for room in self.rooms:
            # set dates
            days = self.daysIterator
            self.weekNum[i].setText(f"Week {self.week}")
            for col in range(7):
                room.setItem(0, col, QTableWidgetItem(self.weeksInYear[days]))
                room.item(0, col).setBackground(QtGui.QColor(220,220,220))
                days += 1
            i += 1
    
    def random_background_color(self):
        '''
        Description: get random background color for the schedule. Makes it 
        easier to read
        '''
        r = random.randint(70,210)
        g = random.randint(100,200)
        b = random.randint(100,220)
        rgb = [r,g,b]
        return (rgb)
    
    def create_schedule(self):
        '''
        Description: set schedule into proper room
        '''
        # set schedules (TODO: need to implement schedules per cohort here)
        
        # row is hours, so index 1 = 8:00AM, index 2 = 8:30AM... 20 = 5:00PM
        # col is days, 0 = Mon, 1 = Tue ... 6 = Sun
        
        # schedule data should include hours as increments of 30 min
        # and should include what weekday they are 
        
        # the ranges here will be replaced with "for row in time" (for row)
        # and "for col in day" (for col)
        
        # wherever a blank is required, a "None" should be present
        
        # 1 = 8:00AM, 2 = 8:30AM... 20 = 5:00PM  
        scheduler = Scheduler()
        if self.cohortUploaded:   
            scheduled_week = scheduler.scheduleCourses(Week(0), self._flattened_cohorts)
                
            for hour in range(1, 20):
                # 0 = Mon, 1 = Tue ... 6 = Sun
                for day in range(7):
                    # self.rooms[0] = 1st room, self.rooms[1] = 2nd room ... etc
                    self.rooms[0].setItem(hour, day, QTableWidgetItem("testScHedule"))
    
    def enable_room(self):
        '''
        Description: gets the room name from file and make the room visible as 
        tab
        '''
        classrooms = getClassrooms()
        tabIndex = 0
        
        # make appropriate tab names according to room num and size
        for classroom in classrooms:
            self.scheduleTab.setTabText(tabIndex, \
                f"{classroom.classRoomNumber} ({classroom.normalCapacity})")
            self.roomNums.append\
                (f"{classroom.classRoomNumber} ({classroom.normalCapacity})")
            tabIndex += 1
        
        # disable every other room tab that is not in use
        for i in range (tabIndex, self.scheduleTab.count()):
            self.scheduleTab.setTabVisible(i, False)
    
    def week_next(self):
        '''
        Description: set schedule to week after current
        '''
        if self.daysIterator >= 359:
            self.daysIterator = 0
        self.daysIterator += 7
        
        if self.week >= 53:
            self.week = 0
        self.week += 1
        
        self.create_dates()
        self.create_schedule()

    def week_prev(self):
        '''
        Description: set schedule to week before current
        '''
        if self.daysIterator <= 0:
            self.daysIterator = 366
        self.daysIterator -= 7
        
        if self.week <= 1:
            self.week = 54
        self.week -= 1
        
        self.create_dates()
        self.create_schedule()

# Initialize The App
def main():
    app = QApplication(sys.argv)
    UIWindow = MainMenu()
    UIWindow.show()
    app.exec()
