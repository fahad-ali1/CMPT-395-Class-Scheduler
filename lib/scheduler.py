"""
Author: Mike Lee, Schuyler Kelly
Purpose: Collection of objects related to weeks and days.
"""
import random
from datetime import datetime, timedelta
from openpyxl.reader.excel import load_workbook
from lib.classrooms import Classroom
from collections import deque
from lib.fileio import getClassrooms
import copy

from collections import deque


class Scheduler:
    def __init__(self):
        self.timeMap = {}
        hours = 8
        minutes = 0
        am_or_pm = "AM"
        for timeValue in range(1, 20):
            self.timeMap[f"{hours}:{minutes:02d} {am_or_pm}"] = timeValue
            hours = hours + 1 if (hours != 12 and minutes == 30) else 1 if (hours == 12 and minutes == 30) else hours
            minutes = 30 if minutes == 0 else 0
            am_or_pm = "AM" if 8 <= hours < 12 else "PM"
            
    def scheduleCourses(self, week, cohorts):
        for cohort in cohorts:
            courseQueue = deque(cohort.programCourses.term1)
            if "BC" in cohort.cohortName or "PC" in cohort.cohortName:
                days = [week.getDay("Monday"), week.getDay("Wednesday")]
            else:
                days = [week.getDay("Tuesday"), week.getDay("Thursday")]

            while courseQueue:
                currentCourse = courseQueue.popleft()
                prefClassroomName = cohort.classroom.classRoomNumber
                lectureLength = currentCourse.lectureLength

                # Determine how many times to schedule the lecture
                if lectureLength == 1.5:
                    schedule_days = days
                else:
                    schedule_days = [random.choice(days)]

                for day in schedule_days:
                
                    # If the preferred classroom is a ghost room, add it to the list of classrooms
                    if prefClassroomName == "??-???" and cohort.classroom not in day.classrooms:
                        day.classrooms.append(cohort.classroom)

                    scheduled_preferred_room = False
                    for i in range(len(day.classrooms)):
                        if (return_value := scheduleLecture(lectureLength, day.classrooms[i].currentBlockTime)):
                            startTime, endTime = return_value
                        else:
                            continue

                        newBlock = timeBlock(startTime, endTime, cohort.cohortName, currentCourse.courseName, 0, 0)
                        if day.classrooms[i].available_at_time(startTime, endTime):
                            if day.classrooms[i].classRoomNumber == prefClassroomName and cohort.size <= day.classrooms[i].normalCapacity:
                                day.classrooms[i].timeBlocks.append(newBlock)
                                day.classrooms[i].currentBlockTime = endTime
                                scheduled_preferred_room = True
                                break

                    if not scheduled_preferred_room:
                        for i in range(len(day.classrooms)):
                            if (return_value := scheduleLecture(lectureLength, day.classrooms[i].currentBlockTime)):
                                startTime, endTime = return_value
                            else:
                                continue

                            newBlock = timeBlock(startTime, endTime, cohort.cohortName, currentCourse.courseName, 0, 0)
                            if day.classrooms[i].available_at_time(startTime, endTime):
                                if cohort.size <= day.classrooms[i].normalCapacity:
                                    day.classrooms[i].timeBlocks.append(newBlock)
                                    day.classrooms[i].currentBlockTime = endTime
                                    break

        return copy.deepcopy(week)


def findAvailableTimeBlock(day, cohort, courseLength):
    for classroom in day.classrooms:
        if classroom.capacity >= cohort.size:
            lastEndTime = '08:00 AM'
            for currentTimeBlock in classroom.timeBlocks:
                if currentTimeBlock.isAvailable(lastEndTime, courseLength, cohort):
                    return classroom, currentTimeBlock
                lastEndTime = currentTimeBlock.endTime
            return classroom, None
    return None, None

def sessionsPerWeek(courseLength):
    if courseLength == 1.5:
        return 2
    else:
        return 1


def createTemplateWeek(classrooms):
    DAYS = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
    templateWeek = Week(0)

    for dayName in DAYS:
        day = Day(dayName)
        day.classrooms = copy.deepcopy(classrooms)
        templateWeek.addDay(day)

    return templateWeek


"""
TIMEBLOCK
---------
ATTRIBUTES
----------
startTime   - course start time
endTime     - course end time
cohortName  - cohort name (string)
courseName  - course name (string)
startDate   - course start date
endDate     - course end date
"""
class timeBlock:
    def __init__(self, startTime, endTime, cohortName, courseName, startDate, endDate):
        self.startTime = startTime
        self.endTime = endTime
        self.cohortName = cohortName
        self.courseName = courseName
        self.startDate = startDate
        self.endDate = endDate
        
    def __repr__(self):
        return f"{self.startTime} - {self.endTime}: Cohort Number <{self.cohortName}>,  Course Name <{self.courseName}>"

    def isAvailable(self, proposed_start_time, proposed_end_time, cohort):
        if self.cohort is not None:
            return False

        startTime = datetime.strptime(self.startTime, '%I:%M %p')
        endTime = datetime.strptime(self.endTime, '%I:%M %p')

        proposed_start_time = datetime.strptime(proposed_start_time, '%I:%M %p')
        proposed_end_time = datetime.strptime(proposed_end_time, '%I:%M %p')

        if proposed_start_time >= startTime and proposed_end_time <= endTime:
            return True
        else:
            return False

    def schedule(self, cohort, course):
        self.cohortName = cohort
        self.courseName = course

"""
DAY
---

ATTRIBUTES
----------
dayName     - name of the day as a string
classrooms   - list of classroom objects

METHODS
-------
getDayName, getTimeBlock - both getters
addTimeBlock - adds a new time block to the day
"""
class Day:
    def __init__(self, dayName):
        self.dayName = dayName # dayName (mon, tues, wed, thurs)
        self.classrooms = []    # list of classroom objects
        
    def __repr__(self):
        string = ""
        for classroom in self.classrooms:
            if classroom.timeBlocks:
                string += f"{classroom.classRoomNumber}:\n"
                for timeBlock in classroom.timeBlocks:
                    string += f"\t\t{timeBlock}\n"
            else:
                string += f"{classroom.classRoomNumber}:\n"
        string.strip("\n")
        return string

    def getDayName(self):
        return self.dayName

    def getClassrooms(self):
        return self.classrooms

    def addClassroom(self, newClassroom):
        self.classrooms.append(newClassroom)


"""
WEEK
----

ATTRIBUTES
----------
weekNumber  - int to represent n/14 weeks
weekDays    - list of day objects

METHODS
-------
getWeekDays - getters
"""
class Week:
    def __init__(self, weekNumber):
        self.weekNumber = weekNumber # int to represent n/14 weeks
        self.days = [       # list of day objects
            Day("Monday"),
            Day("Tuesday"),
            Day("Wednesday"),
            Day("Thursday"),
            Day("Friday"),
            Day("Saturday"),
            Day("Sunday")
        ]
        
        for day in self.days:
            day.classrooms = getClassrooms()
            
    def __repr__(self):
        return f"""Monday:\n{self.days[0]}
Tuesday:\n{self.days[1]}
Wednesday:\n{self.days[2]}
Thursday:\n{self.days[3]}
Friday:\n{self.days[4]}
Saturday:\n{self.days[5]}
Sunday:\n{self.days[6]}"""

    def getWeekNumber(self):
        return self.weekNumber

    def getWeekDays(self):
        return self.days

    def getDay(self, dayName):
        for day in self.days:
            if(day.getDayName() == dayName):
                return day

    def addDay(self, day):
        self.days.append(day)


"""
Purpose: Schedules lecture
Parameters: lecture length, end time
Returns: start time and end time
"""
def scheduleLecture(lectureLength, lastEndTime=None):
    if lectureLength not in [1.5, 2, 3]:
        raise ValueError("Invalid lecture length. Lecture lengths can be 1.5, 2, or 3 hours.")

    startTime = datetime.strptime('08:00:00', '%H:%M:%S')
    endTime = datetime.strptime('17:00:00', '%H:%M:%S')
    lecture_minutes = int(lectureLength * 60)

    if lastEndTime is None:
        startTime = datetime.strptime('08:00:00', '%H:%M:%S')
    else:
        startTime = datetime.strptime(lastEndTime, '%I:%M %p') + timedelta(minutes=10)

    # Round up to the nearest 30-minute interval if the lecture is scheduled within the 10-minute window
    if startTime.minute % 30 != 0:
        minutes_to_round = 30 - (startTime.minute % 30)
        startTime += timedelta(minutes=minutes_to_round)

    if startTime + timedelta(minutes=lecture_minutes) > endTime:
        return False

    endTime = startTime + timedelta(minutes=lecture_minutes) - timedelta(minutes=10)

    return startTime.strftime('%I:%M %p'), endTime.strftime('%I:%M %p')


"""
Purpose: NULL
Paramters: NULL
Returns: NULL
"""
def calculateSessions(transcriptHours, lectureLen):
    pass


def getProgramNumbers(fileName):
    try:
        wb = load_workbook(fileName)
        # wb = load_workbook("data/semester_data.xlsx")
        ws = wb.active
        programNumbersList = []
        for i in range(2, ws.max_row + 1):
            programNumbersList.append([ws['B' + str(i)].value, ws['C' +str(i)].value, ws['D' + str(i)].value])
        return programNumbersList

    except:
        print("Error: File not found")
        return -1


# Expected input: String or None type value from cell F of AllCourses.xlsx
# Expected output: If None type is found then return float indicating default lecture length is returned (1.5 hours)
# Else if value is found then return float indicating minimum lecture length
def getMinimumTime(data):
    if data is None:
        return 1.5
    else:
        return float(data[0])


def checkIntOrDec(num):
    if num.is_integer():
        return int(num)
    else:
        return int(num) + 1


# prototype
# Expected input:
#
#
def calcNumberOfSessions(course):
    testVar = 0.0
    testVar = float(course.getTotalTranscriptHours()) / course.getLectureLength()
    return checkIntOrDec(testVar)


# Expected input: value from cell E of AllCourse.xlsx
# Expected output: returns a string indicating the lab status of a course
def checkIfLab(data):
    if data is None:
        return "Normal"
    elif data == "Lab":
        return "Lab"
    elif data == "Both":
        return "Both"
    elif data == "Virtual":
        return "Virtual"
    elif data == "Online":
        return "Online"


# This function is a help function designed to help with assigning number of sessions to a course
# If the number of course sessions do not divide evenly with the total transcript hours then add an extra hour
# Else if evenly divided, return the number without any change.
# Expected input: a number either an integer or a float
# Expected output returns a number based on if the input had a decimal or not. If no decimal then return number unchanged
# Else if input has a decimal, then return integer value + 1 of input
def checkDecimal(num):
    numToStr = str(num)
    if numToStr.find(".") < 1:
        return int(num) + 1
    else:
        return num


'''
Legend for excel file codes:
SA = Schedule after all classes are done, end of the term
NBOA = Do not schedule immediately before or after other courses
Twice/Half = Class should be schedule twice a week half way through the term
'''
'''
ALLCOURSES.XLSX LEGEND:
A:
Term/course name
B:
Course Description
C:
Total Transcript Hours
D:
Prereqs
E:
Course Type(lab, normal, online/virtual, etc...)
F:
Timeslot
G:
Number of sessions
H:
Order classes should be scheduled(Twice a week halfway through a semester, etc..)
'''
'''
Hours   # of sessiosn
15      10
21      14
35      24
50      34

WHAT TO WORK ON:
return type = string cohort; float start hour; float end hour;
Ideas:

make a class called term object which encapsulates days, weeks.

self.weeky(days)

week object is a list of 7 day objects

The list of weeks which is a copy of the week object above, which changes made for each week.

make a day object



'''
