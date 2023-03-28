"""
Author: Mike Lee
Purpose: Handle file input/output
"""

import csv, re, openpyxl
from openpyxl.reader.excel import load_workbook

# Import local code
from lib.classrooms import Classroom
from lib.courses import Program, Course

"""
Purpose: Reads classroom info from an excel file
Parameters: None
Returns: list of classrooms
"""
def getClassrooms():
    wb = load_workbook('data/AllCourses.xlsx')
    allClassrooms = []
    ws = wb.worksheets[8]
    for i in range(2, ws.max_row + 1):
        if "Lab" in ws['A' + str(i)].value:
            classroom = Classroom(ws['A' + str(i)].value, ws['B' + str(i)].value, isLab=True)
            allClassrooms.append(classroom)
        else:
            classroom = Classroom(ws['A' + str(i)].value, ws['B' + str(i)].value)
            allClassrooms.append(classroom)
    return allClassrooms


# Expected input: value from cell E of AllCourse.xlsx
# Expected output: returns a string indicating the lab status of a course
def checkIfLab(data):
    if data is None:
        return False
    elif data == "Lab":
        return True
    elif data == "Both":
        return "Both"
    elif data == "Virtual":
        return "Virtual"
    elif data == "Online":
        return "Online"
    
# getAllPrograms
# Helper Function
# Purpose: Used to gather a list of program objects representing all available programs at macewan
def getAllPrograms():
    wb = load_workbook('data/AllCourses.xlsx')
    term = None
    allPrograms = []
    for i in range(0, 8):
        ws1 = wb.worksheets[i]
        program = Program(programType=ws1.title)
        for j in range(2, ws1.max_row + 1):
            if re.search('Term [1-3]', ws1['A' + str(j)].value):
                term = ws1['A' + str(j)].value
                pass

            elif ws1['B' + str(j)].value is not None:
                course = Course(courseName=ws1['A' + str(j)].value, courseDescript=ws1['B' + str(j)].value,
                                totalTranscriptHours=ws1['C' + str(j)].value)
                course.setCourseType(str(checkIfLab(ws1['E' + str(j)].value)))
                course.setLectureLength(float(getMinimumTime(ws1['F' + str(j)].value)))
                if ws1['G' + str(j)].value is None:
                    course.setNumberOfSessions("")
                else:
                    course.setNumberOfSessions(ws1['G' + str(j)].value)

                program.addToTerm(course, term)
        allPrograms.append(program)
    return allPrograms

def calculateSessions(transcriptHours, lectureLen):
    pass


# getProgramNumbers
# Helper Function
#
def getProgramNumbers(fileName):
    try:
        wb = load_workbook(fileName)
        # wb = load_workbook("semester_data.xlsx") #fast testing
        ws = wb.active
        programNumbersList = []
        for i in range(2, ws.max_row + 1):
            programNumbersList.append([ws['B' + str(i)].value, ws['C' + str(i)].value, ws['D' + str(i)].value])
        return programNumbersList
    except:
        print("Error: File not found")
        return -1


# Expected input: String or None type value from cell F of Allcourse.xlsx
# Expected output: If None type is found then return float indicating default lecture length is returned (1.5 hours)
# Else if value is found then return float indicating minimum lecture length
def getMinimumTime(data):
    if data is None:
        return 1.5
    else:
        return float(data[0])


def checkIntOrDec(num):
    if num.is_integer():
        return int(num)
    else:
        return int(num) + 1


# prototype
# Expected input:
#
#
def calcNumberOfSessions(course):
    testVar = 0.0
    testVar = float(course.getTotalTranscriptHours()) / course.getLectureLength()
    return checkIntOrDec(testVar)




# This function is a help function designed to help with assigning number of sessions to a course
# If the number of course sessions do not divide evenly with the total transcript hours then add an extra hour
# Else if evenly divided, return the number without any change.
# Expected input: a number either an integer or a float
# Expected output returns a number based on if the input had a decimal or not. If no decimal then return number unchanged
# Else if input has a decimal, then return integer value + 1 of input
def checkDecimal(num):
    numToStr = str(num)
    if numToStr.find(".") < 1:
        return int(num) + 1
    else:
        return num

'''
"""
Purpose: Reads program info from an excel file
Parameters: None
Returns: List of programs
"""
def getAllPrograms():
    wb = load_workbook('data/AllCourses.xlsx')
    term = None
    allPrograms = []
    for i in range(0, 8):
        ws1 = wb.worksheets[i]
        program = Program(programType=ws1.title)
        for j in range(2, ws1.max_row + 1):
            if re.search('Term [1-3]', ws1['A' + str(j)].value):
                term = ws1['A' + str(j)].value
            elif ws1['B' + str(j)].value is not None:
                course = Course(courseName=ws1['A' + str(j)].value, courseDescript=ws1['B' + str(j)].value, \
                                totalTranscriptHours=ws1['C' + str(j)].value)
                program.addToTerm(course, term)
        allPrograms.append(program)
    return allPrograms


'''


"""

Returns a list of lists. Inner list is the number of students per term for each program.
List order is as follows:
BCOM
PCOM
PM
BA
SCMT
BK
FS
DXD
Example: [[1,2,3],[4,5,6],[7,8,9]]
index 0: BCOM term 1,2,3
index 1: PCOM term 1,2,3
index 3: PM term 1,2,3
"""
def getProgramNumbers(fileName):
    try:
        wb = load_workbook(fileName)
        # wb = load_workbook("semester_data.xlsx") #fast testing
        ws = wb.active
        programNumbersList = []
        for i in range(2, ws.max_row + 1):
            programNumbersList.append([ws['B' + str(i)].value, ws['C' + str(i)].value, ws['D' + str(i)].value])
        return programNumbersList
    except:
        # print("Error: File not found")
        return -1

