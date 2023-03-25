"""
Author: Mike Lee, Sankalp Shrivastav
Date: 01/02/2023
Purpose: Rough idea of a course structure
"""

import re
from openpyxl.reader.excel import load_workbook
from datetime import datetime, timedelta


class Course:
    def __init__(self, courseName="", courseDescript="", totalTranscriptHours=0, lectureLength=0, isLab=False):
        self.courseName = courseName
        self.courseDescript = courseDescript
        self.totalTranscriptHours = totalTranscriptHours
        self.lectureLength = 0.0
        self.numberOfSessions = 0
        self.courseType = ""
        self.schedulingInstructions = ""

    def printCourseDetails(self):
        print(self.courseName, " ", self.courseDescript, " ", self.totalTranscriptHours)

    def getEndDate(self):
        # functionality coming soon
        pass

    def getStartDate(self):
        # functionality coming soon
        pass

    def setCourseName(self, courseName):
        self.courseName = courseName

    def setCourseDes(self, courseDes):
        self.courseDescript = courseDes

    def setTranscriptHours(self, transcriptHours):
        self.totalTranscriptHours = transcriptHours

    def setCourseType(self, status):
        self.courseType = status

    def setLectureLength(self, lecLen):
        self.LectureLength = lecLen

    def setScheduleingInstructions(self, instructions):
        self.schedulingInstructions = instructions

    def getSchedulingInstructions(self):
        return self.schedulingInstructions

    def getTotalTranscriptHours(self):
        return self.totalTranscriptHours

    def getCourseName(self):
        return self.courseName

    def getDescription(self):
        return self.courseDescript

    def getCourseType(self):
        return self.courseType

    def getLectureLength(self):
        return self.lectureLength

    def getNumOfSessions(self):
        return self.numberOfSessions

    def getLecTime100(self):
        return self.lectureLength * 100

    def printCourseDetails(self):
        print(self.courseName, " ", self.courseDescript, " ", self.totalTranscriptHours)


class Program:
    def __init__(self, programType=""):
        self.programType = programType
        self.term1 = []
        self.term2 = []
        self.term3 = []

    def setProgram(self, programName):
        self.programType = programName

    def addToTerm(self, course, termNum):
        if termNum == "Term 1":
            self.term1.append(course)
        elif termNum == "Term 2":
            self.term2.append(course)
        elif termNum == "Term 3":
            self.term3.append(course)

    def addProgramType(self, programType):
        self.programType = programType

    def printProgram(self):
        print(self.programType)

    def getAllTerms(self):
        allTerms = [self.term1, self.term2, self.term3]
        return allTerms

    def getTerm1(self):
        return self.term1

    def getTerm2(self):
        return self.term2

    def getTerm3(self):
        return self.term3


class Schedule:
    def __init__(self, day="", courseName="", timeSlot="", cohortName=""):
        self.day = day
        self.courseName = courseName
        self.timeSlot = timeSlot
        self.cohortName = cohortName
        self.currentCapacity = 0


class ScheduleNode:
    def __init__(self, time, cohort, start_date, end_date):
        self.time = time
        self.cohort = cohort
        self.start_date = start_date
        self.end_date = end_date
        self.prev = None
        self.next = None


def getCourse(fileName, courseName):

    wb = load_workbook(fileName)

    for sheet in range(8):

        ws = wb.worksheets[sheet]
        
        for row in range(2, ws.max_row + 1):

            #look for the course name if found return the location [row, col]

            if re.search("^Term", ws["A" + str(row)].value): continue
            # print(ws["A" + str(row)].value)
            if re.search(courseName, ws["A" + str(row)].value): return [sheet, row]

    return None


def changeCourseInfo(courseLocation, courseInfo, fileName):

    sheet, cRow = courseLocation[0], courseLocation[1]
    wb = load_workbook(fileName)

    ws = wb.worksheets[sheet]
    count = 0
    
    for cells in ws[str(cRow)]:
        #swap the info from courseInfo

        cells.value = courseInfo[count]
        # print(cells.value, end="\t") #comment this when demo is done
        count += 1
    wb.save(fileName) #saving the edited file as 

