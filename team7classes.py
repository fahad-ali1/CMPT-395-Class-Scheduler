"""
Author: Mike Lee
ID: 3067768
Date: 01/02/2023
Purpose: Collection of classes and methods needed for the project
"""

import re

from openpyxl.reader.excel import load_workbook
from datetime import datetime, timedelta


# Course Class:
# Represents a course
# Attributes include:
# courseName = name of the course
# courseDescript = course description
# lectureLength = length of each lecture
# courseType = the type of each course(normal, online/virtual, lab)
# schedulingInstructions a string indicating how to schedule a course, for example: NBOA = do not schedule before or
# after another course
# numberOfSessions = a string indicating n number of sessions. getNumOfSessions will return just the number
class Course:
    def __init__(self, courseName="", courseDescript="", totalTranscriptHours=0):
        self.courseName = courseName
        self.courseDescript = courseDescript
        self.totalTranscriptHours = totalTranscriptHours
        self.lectureLength = 0.0
        self.numberOfSessions = 0
        self.courseType = ""
        self.schedulingInstructions = ""

    def setCourseName(self, courseName):
        self.courseName = courseName

    def setCourseDes(self, courseDes):
        self.courseDescript = courseDes

    def setNumberOfSessions(self, sessions):
        self.numberOfSessions = sessions

    def setTranscriptHours(self, transcriptHours):
        self.totalTranscriptHours = transcriptHours

    def setCourseType(self, status):
        self.courseType = status

    def setLectureLength(self, lecLen):
        self.lectureLength = lecLen

    def setschedulingInstructions(self, instructions):
        self.schedulingInstructions = instructions

    def getschedulingInstructions(self):
        return self.schedulingInstructions

    def getTotalTranscriptHours(self):
        return self.totalTranscriptHours

    def getcourseName(self):
        return self.courseName

    def getdescription(self):
        return self.courseDescript

    def getcourseType(self):
        return self.courseType

    def getLectureLength(self):
        return self.lectureLength

    def getNumOfSessions(self):
        return self.numberOfSessions

    def getLecTime100(self):
        return self.lectureLength * 100

    def printCourseDetails(self):
        print(self.courseName, " ", self.courseDescript, " ", self.totalTranscriptHours)


# Program:
# Represents a program at macewan
# Attributes include:
# programType = the type(name) of the program
# term1 = list of all courses(course class) for term 1
# term2 = list of all courses(course class) for term 2
# term3 = list of all courses(course class) for term 3
class Program:
    def __init__(self, programType=""):
        self.programType = programType
        self.term1 = []
        self.term2 = []
        self.term3 = []

    def setProgram(self, programName):
        self.programType = programName

    def addToTerm(self, course, termNum):
        termNum = termNum.strip()
        if termNum == "Term 1":
            self.term1.append(course)
        elif termNum == "Term 2":
            self.term2.append(course)
        elif termNum == "Term 3":
            self.term3.append(course)

    def addProgramType(self, programType):
        self.programType = programType

    def printProgram(self):
        print(self.programType)

    def getAllTerms(self):
        allTerms = [self.term1, self.term2, self.term3]
        return allTerms

    def getTerm1(self):
        return self.term1

    def getTerm2(self):
        return self.term2

    def getTerm3(self):
        return self.term3


# Classroom:
# Represents a physical classroom
# Attributes include:
# classRoomNumber = classroom number/name
# normalCapacity = capacity
# lab = is it a lab
# isGhost = is it a ghost room (room that would be needed for extra students)
# schedule = schedule linked list representing the schedule
class Classroom:
    def __init__(self, cohort, classRoomNumber='', normalCapacity=0, lab=False):
        self.classRoomNumber = classRoomNumber
        self.normalCapacity = normalCapacity
        self.lab = lab
        self.isGhost = False
        self.schedule = None
        self.currentStudents = 0  # will make a setter for it
        self.inUse = False
        self.cohort = cohort

    #
    def setGhost(self):
        self.isGhost = True

    def setSchedule(self, schedule):
        self.schedule = schedule

    def printClassroom(self):
        print("Classroom #: ", self.classRoomNumber, "Normal Capacity: ", self.normalCapacity,
              "lab: ", self.lab, " isGhost: ", self.isGhost)


# Cohort Class takes in a cohort name and size initially
# Methods will be created to add a programs to mainProgramCourses
# and electiveProgramCourses

class Cohort:
    def __init__(self, classroom, cohortName="", size=0):
        self.cohortName = cohortName
        self.size = size
        self.ProgramCourses = None  # program class
        self.prereq = {}
        self.currentCourse = None
        self.classroom = classroom

    def setmainProgram(self, program):
        self.ProgramCourses = program

    def setClassrom(self, classroom):
        self.classroom = classroom

    def getCohortName(self):
        return self.cohortName

    def getMain(self):
        return self.mainProgramCourses

# time block object.
# Attributes: startTime(dateTime hour), endTime(dateTime hour), cohortName(string), startDate(dateTime date),
# endTime(dateTime date)

class timeBlock:
    def __init__(self, startTime, endTime, cohortName, courseName, startDate, endDate):
        self.startTime = startTime  # course start time
        self.endTime = endTime  # course end time
        self.cohortName = cohortName  # cohort name (string)
        self.courseName = courseName  # course name (string)
        self.startDate = startDate  # courses start date
        self.endDate = endDate  # courses end date

# day object.
# Attributes: dayName(string), timeBlocks(list of time block objects)

class day:
    def __init__(self, dayName):
        self.dayName = dayName  # day name (mon,wed,tues,thurs)
        self.timeBlocks = []  # list of time block objects see above empty to start with

    def getDayName(self):
        return self.dayName

    def getTimeBlocks(self):
        return self.timeBlocks

    def addTimeBlock(self, newTimeBlock):
        self.timeBlocks.append(newTimeBlock)


# week object.
# Attributes: weekNumber(integer), weekDays(list of day objects)

class week:
    def __init__(self, weekNumber, weekDays):
        self.weekNumber = weekNumber  # int to represent n/14 weeks
        self.weekDays = weekDays  # list of day objects

    def getWeekDays(self):
        return self.weekDays  # list of day objects


def scheduleLecture(lectureLength, lastEndTime=None):
    if lectureLength not in [1.5, 2, 3]:
        raise ValueError("Invalid lecture length. Lecture lengths can be 1.5, 2, or 3 hours.")

    startTime = datetime.strptime('08:00:00', '%H:%M:%S')
    endTime = datetime.strptime('17:00:00', '%H:%M:%S')
    lecture_minutes = int(lectureLength * 60)

    if lastEndTime is None:
        startTime = datetime.strptime('08:00:00', '%H:%M:%S')
    else:
        startTime = datetime.strptime(lastEndTime, '%I:%M %p') + timedelta(minutes=10)

    # Round up to the nearest 30-minute interval if the lecture is scheduled within the 10-minute window
    if startTime.minute % 30 != 0:
        minutes_to_round = 30 - (startTime.minute % 30)
        startTime += timedelta(minutes=minutes_to_round)

    if startTime + timedelta(minutes=lecture_minutes) > endTime:
        raise ValueError("Course cannot be booked after 5pm.")

    endTime = startTime + timedelta(minutes=lecture_minutes) - timedelta(minutes=10)

    return startTime.strftime('%I:%M %p'), endTime.strftime('%I:%M %p')

# need a function to get start and end dates in 2400 hour format
# start/end is for schedule node

# getClassrooms:
# Helper Fucntion
# Purpose: Used to gather list of classroom objects representing all available classrooms
def getClassrooms():
    wb = load_workbook('AllCourses.xlsx')
    allClassrooms = []
    ws = wb.worksheets[8]
    for i in range(2, ws.max_row + 1):
        if "Lab" in ws['A' + str(i)].value:
            classroom = Classroom(ws['A' + str(i)].value, ws['B' + str(i)].value, True)
            allClassrooms.append(classroom)
        else:
            classroom = Classroom(ws['A' + str(i)].value, ws['B' + str(i)].value)
            allClassrooms.append(classroom)
    return allClassrooms


# getAllPrograms
# Helper Function
# Purpose: Used to gather a list of program objects representing all available programs at macewan
def getAllPrograms():
    wb = load_workbook('AllCourses.xlsx')
    term = None
    allPrograms = []
    for i in range(0, 8):
        ws1 = wb.worksheets[i]
        program = Program(programType=ws1.title)
        for j in range(2, ws1.max_row + 1):
            if re.search('Term [1-3]', ws1['A' + str(j)].value):
                term = ws1['A' + str(j)].value
                pass

            elif ws1['B' + str(j)].value is not None:
                course = Course(courseName=ws1['A' + str(j)].value, courseDescript=ws1['B' + str(j)].value,
                                totalTranscriptHours=ws1['C' + str(j)].value)
                course.setCourseType(str(checkIfLab(ws1['E' + str(j)].value)))
                course.setLectureLength(getMinimumTime(ws1['F' + str(j)].value))
                if ws1['G' + str(j)].value is None:
                    course.setNumberOfSessions(calcNumberOfSessions(course))
                else:
                    course.setNumberOfSessions(ws1['G' + str(j)].value)

                program.addToTerm(course, term)
        allPrograms.append(program)
    return allPrograms


def calculateSessions(transcriptHours, lectureLen):
    pass


# getProgramNumbers
# Helper Function
#
def getProgramNumbers(fileName):
    try:
        wb = load_workbook(fileName)
        # wb = load_workbook("semester_data.xlsx") #fast testing
        ws = wb.active
        programNumbersList = []
        for i in range(2, ws.max_row + 1):
            programNumbersList.append([ws['B' + str(i)].value, ws['C' + str(i)].value, ws['D' + str(i)].value])
        return programNumbersList
    except:
        print("Error: File not found")
        return -1


# Expected input: String or None type value from cell F of Allcourse.xlsx
# Expected output: If None type is found then return float indicating default lecture length is returned (1.5 hours)
# Else if value is found then return float indicating minimum lecture length
def getMinimumTime(data):
    if data is None:
        return 1.5
    else:
        return float(data[0])


def checkIntOrDec(num):
    if num.is_integer():
        return int(num)
    else:
        return int(num) + 1


# prototype
# Expected input:
#
#
def calcNumberOfSessions(course):
    testVar = 0.0
    testVar = float(course.getTotalTranscriptHours()) / course.getLectureLength()
    return checkIntOrDec(testVar)


# Expected input: value from cell E of AllCourse.xlsx
# Expected output: returns a string indicating the lab status of a course
def checkIfLab(data):
    if data is None:
        return "Normal"
    elif data == "Lab":
        return "Lab"
    elif data == "Both":
        return "Both"
    elif data == "Virtual":
        return "Virtual"
    elif data == "Online":
        return "Online"


# This function is a help function designed to help with assigning number of sessions to a course
# If the number of course sessions do not divide evenly with the total transcript hours then add an extra hour
# Else if evenly divided, return the number without any change.
# Expected input: a number either an integer or a float
# Expected output returns a number based on if the input had a decimal or not. If no decimal then return number unchanged
# Else if input has a decimal, then return integer value + 1 of input
def checkDecimal(num):
    numToStr = str(num)
    if numToStr.find(".") < 1:
        return int(num) + 1
    else:
        return num


'''
Legend for excel file codes:
SA = Schedule after all classes are done, end of the term
NBOA = Do not schedule immediately before or after other courses
Twice/Half = Class should be schedule twice a week half way through the term
'''
'''
ALLCOURSES.XLSX LEGEND:
A:
Term/course name
B:
Course Description
C:
Total Transcript Hours
D:
Prereqs
E:
Course Type(lab, normal, online/virtual, etc...)
F:
Timeslot
G:
Number of sessions
H:
Order classes should be scheduled(Twice a week halfway through a semester, etc..)
'''
'''
Hours   # of sessiosn
15      10
21      14
35      24
50      34

WHAT TO WORK ON:
return type = string cohort; float start hour; float end hour;
Ideas:

make a class called term object which encapsulates days, weeks.

self.weeky(days)

week object is a list of 7 day objects

The list of weeks which is a copy of the week object above, which changes made for each week.

make a day object



'''
